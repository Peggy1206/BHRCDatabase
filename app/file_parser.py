import asyncio
import io
import base64
import httpx
import anthropic
import trafilatura
from docx import Document
from openpyxl import load_workbook
from app.config import settings

claude = anthropic.Anthropic(api_key=settings.anthropic_api_key)


_BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
}


async def fetch_url_content(url: str) -> tuple[str, str]:
    """Fetch and extract clean text from a URL. Returns (text, final_url)."""
    async with httpx.AsyncClient(follow_redirects=True, timeout=15, headers=_BROWSER_HEADERS) as client:
        response = await client.get(url)
        final_url = str(response.url)
        html = response.text

    text = trafilatura.extract(html, include_comments=False, include_tables=True)
    if not text:
        text = html[:5000]
    return text, final_url


async def parse_image_bytes(image_bytes: bytes) -> str:
    """Use Claude Vision to describe image content."""
    encoded = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = await asyncio.to_thread(
        claude.messages.create,
        model="claude-sonnet-4-6",
        max_tokens=1000,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/jpeg", "data": encoded},
                    },
                    {
                        "type": "text",
                        "text": (
                            "Please describe this image in detail. "
                            "If it contains text, extract all readable text. "
                            "Focus on information relevant to business, talent, or industry trends."
                        ),
                    },
                ],
            }
        ],
    )
    return response.content[0].text


def _parse_docx(file_bytes: bytes) -> str:
    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                lines.append(row_text)
    return "\n".join(lines)


def parse_document_bytes(file_bytes: bytes, filename: str) -> str:
    """Parse Word or Excel file bytes into plain text."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("docx", "doc"):
        return _parse_docx(file_bytes)
    if ext in ("xlsx", "xls"):
        return _parse_xlsx(file_bytes)
    try:
        return file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return file_bytes.decode("latin-1", errors="replace")
